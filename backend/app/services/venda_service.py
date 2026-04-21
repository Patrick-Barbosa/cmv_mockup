from datetime import date, datetime
from io import BytesIO

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import func, select, delete, insert, and_
from sqlalchemy.dialects.postgresql import insert as pg_insert

from backend.app.database.models import Produto, Venda
from backend.app.schemas.venda import VendaImportRowModel, BulkImportVendasModel, ImportStrategy

EXPECTED_UPLOAD_COLUMNS = [
# ... (rest of imports and class start)
    "data",
    "id_loja",
    "id_produto",
    "quantidade_produto",
    "valor_total",
]


class VendaService:
    def __init__(self, session):
        self.session = session

    @staticmethod
    def _month_bounds(month: str) -> tuple[date, date]:
        month_start = datetime.strptime(month, "%Y-%m").date().replace(day=1)
        if month_start.month == 12:
            next_month = date(month_start.year + 1, 1, 1)
        else:
            next_month = date(month_start.year, month_start.month + 1, 1)
        return month_start, next_month

    @staticmethod
    def _to_float(value):
        return float(value) if value is not None else 0.0

    @staticmethod
    def _ideal_unit_cost(produto_tipo: str | None, custo: float | None, preco_referencia: float | None) -> float | None:
        if produto_tipo == "insumo":
            if preco_referencia is not None:
                return float(preco_referencia)
            if custo is not None:
                return float(custo)
            return None

        if custo is None:
            return None
        return float(custo)

    async def import_excel(self, file_bytes: bytes):
        try:
            workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Não foi possível ler o arquivo Excel. Use um arquivo .xlsx válido.") from exc

        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = list(header_row or [])

        if headers != EXPECTED_UPLOAD_COLUMNS:
            workbook.close()
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "Cabeçalho inválido. As colunas devem bater exatamente com o formato esperado.",
                    "expected_columns": EXPECTED_UPLOAD_COLUMNS,
                    "received_columns": headers,
                },
            )

        vendas_validas: list[VendaImportRowModel] = []
        erros: list[str] = []

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row_values = list(row[: len(EXPECTED_UPLOAD_COLUMNS)])
            if len(row_values) < len(EXPECTED_UPLOAD_COLUMNS):
                row_values.extend([None] * (len(EXPECTED_UPLOAD_COLUMNS) - len(row_values)))

            if all(value in (None, "") for value in row_values):
                continue

            row_data = dict(zip(EXPECTED_UPLOAD_COLUMNS, row_values))
            try:
                vendas_validas.append(VendaImportRowModel.model_validate(row_data))
            except Exception as exc:
                erros.append(f"Linha {row_number}: {exc}")

        workbook.close()

        if erros:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "Arquivo inválido. Corrija os erros antes de importar.",
                    "errors": erros,
                },
            )

        if not vendas_validas:
            raise HTTPException(status_code=422, detail="Nenhuma linha válida encontrada para importação.")

        self.session.add_all([
            Venda(
                data=venda.data,
                id_loja=venda.id_loja,
                id_produto=venda.id_produto,
                quantidade_produto=venda.quantidade_produto,
                valor_total=venda.valor_total,
            )
            for venda in vendas_validas
        ])
        await self.session.commit()

        filtros = await self.get_filters()
        return {
            "message": "Vendas importadas com sucesso.",
            "linhas_importadas": len(vendas_validas),
            **filtros,
        }

    async def get_filters(self):
        month_label = func.to_char(Venda.data, 'YYYY-MM').label('mes')
        months_subquery = (
            select(month_label)
            .distinct()
            .subquery()
        )

        lojas_result = await self.session.execute(
            select(Venda.id_loja).distinct().order_by(Venda.id_loja)
        )
        meses_result = await self.session.execute(
            select(months_subquery.c.mes)
            .order_by(months_subquery.c.mes.desc())
        )

        return {
            "lojas": lojas_result.scalars().all(),
            "meses": meses_result.scalars().all(),
        }

    async def get_store_month_analysis(self, store_id: str, month: str):
        month_start, next_month = self._month_bounds(month)

        sales_subquery = (
            select(
                Venda.id_produto.label("id_produto"),
                func.sum(Venda.quantidade_produto).label("quantidade_total"),
                func.sum(Venda.valor_total).label("valor_total"),
            )
            .where(
                Venda.id_loja == store_id,
                Venda.data >= month_start,
                Venda.data < next_month,
            )
            .group_by(Venda.id_produto)
            .subquery()
        )

        result = await self.session.execute(
            select(
                sales_subquery.c.id_produto,
                sales_subquery.c.quantidade_total,
                sales_subquery.c.valor_total,
                Produto.id.label("produto_id_interno"),
                Produto.nome.label("produto_nome"),
                Produto.tipo.label("produto_tipo"),
                Produto.id_produto_externo,
                Produto.custo.label("custo_unitario_ideal"),
                Produto.preco_referencia.label("preco_referencia"),
            )
            .select_from(
                sales_subquery.outerjoin(
                    Produto,
                    Produto.id_produto_externo == sales_subquery.c.id_produto,
                )
            )
            .order_by(sales_subquery.c.valor_total.desc(), sales_subquery.c.id_produto)
        )

        produtos = []
        receita_total = 0.0
        receita_vinculada = 0.0
        receita_sem_vinculo = 0.0
        custo_ideal_total = 0.0
        quantidade_total = 0
        produtos_vinculados = 0
        produtos_sem_vinculo = 0

        for row in result.all():
            quantidade = int(row.quantidade_total or 0)
            valor_total = self._to_float(row.valor_total)
            preco_medio = valor_total / quantidade if quantidade > 0 else None
            custo_unitario_ideal = self._ideal_unit_cost(
                row.produto_tipo,
                row.custo_unitario_ideal,
                row.preco_referencia,
            )
            custo_ideal_produto = custo_unitario_ideal * quantidade if custo_unitario_ideal is not None else None
            cmv_ideal_percentual = (
                (custo_unitario_ideal / preco_medio) * 100
                if custo_unitario_ideal is not None and preco_medio and preco_medio > 0
                else None
            )
            vinculado = row.produto_id_interno is not None

            receita_total += valor_total
            quantidade_total += quantidade

            if vinculado:
                produtos_vinculados += 1
                receita_vinculada += valor_total
                custo_ideal_total += custo_ideal_produto or 0.0
            else:
                produtos_sem_vinculo += 1
                receita_sem_vinculo += valor_total

            produtos.append({
                "id_produto": row.id_produto,
                "produto_id_interno": row.produto_id_interno,
                "produto_nome": row.produto_nome,
                "produto_tipo": row.produto_tipo,
                "id_produto_externo": row.id_produto_externo,
                "vinculado": vinculado,
                "quantidade_total": quantidade,
                "valor_total": valor_total,
                "preco_medio": preco_medio,
                "custo_unitario_ideal": custo_unitario_ideal,
                "custo_ideal_total": custo_ideal_produto,
                "cmv_ideal_percentual": cmv_ideal_percentual,
            })

        cmv_ideal_percentual = (
            (custo_ideal_total / receita_vinculada) * 100
            if receita_vinculada > 0
            else None
        )

        return {
            "loja_id": store_id,
            "mes": month,
            "resumo": {
                "receita_total": receita_total,
                "receita_vinculada": receita_vinculada,
                "receita_sem_vinculo": receita_sem_vinculo,
                "custo_ideal_total": custo_ideal_total,
                "cmv_ideal_percentual": cmv_ideal_percentual,
                "quantidade_total": quantidade_total,
                "produtos_vinculados": produtos_vinculados,
                "produtos_sem_vinculo": produtos_sem_vinculo,
            },
            "produtos": produtos,
        }

    async def get_product_monthly_analysis(self, produto_id: int):
        produto_result = await self.session.execute(select(Produto).where(Produto.id == produto_id))
        produto = produto_result.scalar_one_or_none()
        if not produto:
            raise HTTPException(status_code=404, detail=f"Produto com id={produto_id} não encontrado.")

        if not produto.id_produto_externo:
            return {
                "produto": {
                    "id": produto.id,
                    "nome": produto.nome,
                    "tipo": produto.tipo,
                    "id_produto_externo": produto.id_produto_externo,
                    "custo_unitario_ideal": self._ideal_unit_cost(produto.tipo, produto.custo, produto.preco_referencia),
                },
                "possui_vinculo_externo": False,
                "linhas": [],
            }

        mes_expr = func.to_char(Venda.data, 'YYYY-MM').label("mes")
        result = await self.session.execute(
            select(
                mes_expr,
                Venda.id_loja.label("loja_id"),
                func.sum(Venda.quantidade_produto).label("quantidade_total"),
                func.sum(Venda.valor_total).label("valor_total"),
            )
            .where(Venda.id_produto == produto.id_produto_externo)
            .group_by(mes_expr, Venda.id_loja)
            .order_by(mes_expr.desc(), Venda.id_loja)
        )

        linhas = []
        for row in result.all():
            quantidade = int(row.quantidade_total or 0)
            valor_total = self._to_float(row.valor_total)
            preco_medio = valor_total / quantidade if quantidade > 0 else None
            custo_unitario_ideal = self._ideal_unit_cost(produto.tipo, produto.custo, produto.preco_referencia)
            custo_ideal_total = custo_unitario_ideal * quantidade if custo_unitario_ideal is not None else None
            cmv_ideal_percentual = (
                (custo_unitario_ideal / preco_medio) * 100
                if custo_unitario_ideal is not None and preco_medio and preco_medio > 0
                else None
            )
            linhas.append({
                "mes": row.mes,
                "loja_id": row.loja_id,
                "quantidade_total": quantidade,
                "valor_total": valor_total,
                "preco_medio": preco_medio,
                "custo_unitario_ideal": custo_unitario_ideal,
                "custo_ideal_total": custo_ideal_total,
                "cmv_ideal_percentual": cmv_ideal_percentual,
            })

        return {
            "produto": {
                "id": produto.id,
                "nome": produto.nome,
                "tipo": produto.tipo,
                "id_produto_externo": produto.id_produto_externo,
                "custo_unitario_ideal": self._ideal_unit_cost(produto.tipo, produto.custo, produto.preco_referencia),
            },
            "possui_vinculo_externo": True,
            "linhas": linhas,
        }

    async def bulk_import(self, payload: BulkImportVendasModel):
        if not payload.rows:
            return {"message": "Nenhuma linha para importar.", "linhas_importadas": 0}

        if payload.strategy == ImportStrategy.OVERWRITE:
            # Delete combinations of (data, id_loja) present in the payload
            # Replace exactly what's being sent for those days/stores
            days_lojas = {(row.data, row.id_loja) for row in payload.rows}
            for d, l in days_lojas:
                await self.session.execute(
                    delete(Venda).where(and_(Venda.data == d, Venda.id_loja == l))
                )

        # Prepare for insertion
        insert_data = [
            {
                "data": row.data,
                "id_loja": row.id_loja,
                "id_produto": row.id_produto,
                "quantidade_produto": row.quantidade_produto,
                "valor_total": row.valor_total,
            }
            for row in payload.rows
        ]

        stmt = pg_insert(Venda).values(insert_data)
        
        if payload.strategy == ImportStrategy.APPEND:
            stmt = stmt.on_conflict_do_nothing(constraint='uq_venda_data_loja_produto')
        else:
            # For overwrite strategy, we already deleted, but if the payload itself 
            # has internal duplicates (though frontend aggregates), handle with update
            stmt = stmt.on_conflict_do_update(
                constraint='uq_venda_data_loja_produto',
                set_={
                    "quantidade_produto": stmt.excluded.quantidade_produto,
                    "valor_total": stmt.excluded.valor_total,
                }
            )

        await self.session.execute(stmt)
        await self.session.commit()

        filtros = await self.get_filters()
        return {
            "message": "Vendas importadas com sucesso.",
            "linhas_importadas": len(payload.rows),
            **filtros,
        }

    async def get_missing_skus(self, page: int = 1, size: int = 50):
        # Base query to identify SKUs in Venda that are NOT in Produto.id_produto_externo
        # Using a subquery or outer join to find missing links
        
        # 1. First, get total count of unique missing SKUs
        count_stmt = (
            select(func.count(func.distinct(Venda.id_produto)))
            .select_from(Venda)
            .outerjoin(Produto, Venda.id_produto == Produto.id_produto_externo)
            .where(Produto.id.is_(None))
        )
        total_count_result = await self.session.execute(count_stmt)
        total_count = total_count_result.scalar() or 0

        # 2. Get the paginated results
        stmt = (
            select(
                Venda.id_produto.label("id_produto_externo"),
                func.sum(Venda.quantidade_produto).label("quantidade_total"),
                func.sum(Venda.valor_total).label("valor_total"),
                func.count(Venda.id).label("vendas_count")
            )
            .outerjoin(Produto, Venda.id_produto == Produto.id_produto_externo)
            .where(Produto.id.is_(None))
            .group_by(Venda.id_produto)
            .order_by(func.sum(Venda.valor_total).desc())
            .limit(size)
            .offset((page - 1) * size)
        )
        
        result = await self.session.execute(stmt)
        items = []
        for row in result.all():
            items.append({
                "id_produto_externo": row.id_produto_externo,
                "quantidade_total": int(row.quantidade_total),
                "valor_total": float(row.valor_total),
                "vendas_count": int(row.vendas_count)
            })

        import math
        pages = math.ceil(total_count / size) if size > 0 else 0

        return {
            "total": total_count,
            "page": page,
            "size": size,
            "pages": pages,
            "items": items
        }
